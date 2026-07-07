from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExportService:
    def export_base_history(self, result):
        wb = Workbook()

        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        ws_registros = wb.create_sheet("Registros Base")

        self._build_summary_sheet(ws_resumen, result)
        self._build_records_sheet(ws_registros, result)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _build_summary_sheet(self, ws, result):
        ws["A1"] = "BASE HISTÓRICA INTELIGENTE - CDA IA"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:D1")

        aspecto = result.get("aspecto") or {}
        nivel = result.get("nivel_usado") or {}
        estadisticas = result.get("estadisticas") or {}
        reales = estadisticas.get("reales") or {}
        depuradas = estadisticas.get("depuradas") or {}
        recomendacion = estadisticas.get("recomendacion") or {}
        filters = result.get("filters") or {}

        ia = result.get("ia") or {}
        ia_confianza = ia.get("confianza") or {}
        ia_rango = ia.get("rango_ideal") or {}

        decision_ia = result.get("decision_ia") or {}

        rows = [
            ("Fuente", result.get("source")),
            ("Marca solicitada", filters.get("marca")),
            ("Línea solicitada", filters.get("linea")),
            ("Modelo solicitado", filters.get("anio_modelo")),
            ("Código histórico", filters.get("codigo_historico")),
            ("", ""),
            ("Código concat", aspecto.get("codigo_concat")),
            ("Nombre aspecto", aspecto.get("nombre_aspecto")),
            ("Nombre grupo", aspecto.get("nombre_grupo")),
            ("Grupo revisión", aspecto.get("grupo_revision")),
            ("Unidad medida", aspecto.get("unidad_medida")),
            ("Tipo falla", aspecto.get("tipo_falla")),
            ("Código norma", aspecto.get("codigo_norma")),
            ("", ""),
            ("Nivel usado", nivel.get("nivel")),
            ("Descripción nivel", nivel.get("descripcion")),
            ("Cantidad encontrada", nivel.get("cantidad_encontrada")),
            ("Mínimo requerido", nivel.get("minimo_requerido")),
            ("", ""),
            ("Valor recomendado", recomendacion.get("valor_referencia")),
            ("Método recomendación", recomendacion.get("metodo")),
            ("Motivo", recomendacion.get("motivo")),
            ("", ""),
            ("CRITERIO IA", ""),
            ("Valor objetivo IA", ia.get("valor_objetivo")),
            ("Rango ideal mínimo", ia_rango.get("min")),
            ("Rango ideal máximo", ia_rango.get("max")),
            ("Confianza IA", ia_confianza.get("nivel")),
            ("Score IA", ia_confianza.get("score")),
            ("Estrellas", ia_confianza.get("estrellas")),
            ("Método IA", ia.get("metodo")),
            ("Criterio IA", ia.get("criterio")),
            ("", ""),
            ("DECISIÓN IA FINAL", ""),
            ("Valor final IA", decision_ia.get("valor_final")),
            ("Histórico IA", decision_ia.get("historico")),
            ("CatBoost IA", decision_ia.get("catboost")),
            ("Diferencia IA", decision_ia.get("diferencia")),
            ("Método decisión", decision_ia.get("metodo")),
            ("Peso histórico", decision_ia.get("peso_historico")),
            ("Peso CatBoost", decision_ia.get("peso_catboost")),
            ("Interpretación decisión", decision_ia.get("interpretacion")),
            ("Cantidad real", reales.get("cantidad")),
            ("Min real", reales.get("min")),
            ("Max real", reales.get("max")),
            ("Promedio real", reales.get("promedio")),
            ("Mediana real", reales.get("mediana")),
            ("P5", reales.get("p5")),
            ("P95", reales.get("p95")),
            ("", ""),
            ("Cantidad depurada", depuradas.get("cantidad")),
            ("Min depurado", depuradas.get("min")),
            ("Max depurado", depuradas.get("max")),
            ("Promedio depurado", depuradas.get("promedio")),
            ("Mediana depurada", depuradas.get("mediana")),
        ]

        thin = Side(style="thin", color="D9EAD3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_index, (label, value) in enumerate(rows, start=3):
            label_cell = ws.cell(row=row_index, column=1, value=label)
            value_cell = ws.cell(row=row_index, column=2, value=value)

            label_cell.border = border
            value_cell.border = border

            label_cell.alignment = Alignment(vertical="top")
            value_cell.alignment = Alignment(wrap_text=True, vertical="top")

            if label:
                label_cell.font = Font(bold=True)

            if label in ["CRITERIO IA", "DECISIÓN IA FINAL"]:
                label_cell.fill = PatternFill("solid", fgColor="006100")
                label_cell.font = Font(bold=True, color="FFFFFF")
                value_cell.fill = PatternFill("solid", fgColor="006100")

            if label in [
                "Valor objetivo IA",
                "Rango ideal mínimo",
                "Rango ideal máximo",
                "Confianza IA",
                "Score IA",
                "Estrellas",
                "Valor final IA",
                "Histórico IA",
                "CatBoost IA",
                "Diferencia IA",
                "Peso histórico",
                "Peso CatBoost",
            ]:
                label_cell.fill = PatternFill("solid", fgColor="E2F0D9")
                value_cell.fill = PatternFill("solid", fgColor="E2F0D9")

            if label == "Valor objetivo IA":
                value_cell.font = Font(bold=True, size=14, color="006100")

            if label in ["Valor objetivo IA", "Valor final IA"]:
                value_cell.font = Font(bold=True, size=14, color="006100")    

            if label == "Score IA" and value_cell.value is not None:
                value_cell.value = f"{value_cell.value} / 100"

        ws.freeze_panes = "A3"
        self._auto_width(ws)

    def _build_records_sheet(self, ws, result):
        headers = [
            "CDA",
            "Año proceso",
            "OT",
            "Fecha proceso",
            "Placa",
            "Marca",
            "Línea",
            "Modelo",
            "Código histórico",
            "Valor medición",
            "Valor norma",
            "Tipo línea",
            "Tipo servicio",
            "Combustible",
            "Peso bruto",
            "Aprobación",
        ]

        thin = Side(style="thin", color="D9EAD3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._style_header(cell)
            cell.border = border

        registros = result.get("registros") or []

        if not registros:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            cell = ws.cell(
                row=2,
                column=1,
                value="No se encontraron registros base para los filtros aplicados.",
            )
            cell.font = Font(italic=True, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            self._auto_width(ws)
            return

        for row_index, item in enumerate(registros, start=2):
            values = [
                self._get_value(item, "CDA"),
                self._get_value(item, "ANIO_PROCESO", "anio_proceso"),
                self._get_value(item, "OT"),
                self._get_value(item, "F_PROCESO", "fecha_proceso"),
                self._get_value(item, "PLACA", "placa"),
                self._get_value(item, "MARCA", "marca"),
                self._get_value(item, "LINEA", "linea"),
                self._get_value(item, "ANIO_MODELO", "anio_modelo", "Año Modelo"),
                self._get_value(
                    item,
                    "CODIGO_HISTORICO",
                    "codigo_historico",
                    "COD. HISTORICO",
                    "codigo_concat",
                ),
                self._get_value(
                    item,
                    "VALOR_MEDICION",
                    "valor_medicion",
                    "Valor Medición",
                ),
                self._get_value(
                    item,
                    "VALOR_NORMA",
                    "valor_norma",
                    "Valor Norma",
                ),
                self._get_value(
                    item,
                    "TIPO_LINEA",
                    "tipo_linea",
                    "Tipo Línea",
                ),
                self._get_value(
                    item,
                    "TIPO_SERVICIO",
                    "tipo_servicio",
                    "Tipo Servicio",
                ),
                self._get_value(
                    item,
                    "COMBUSTIBLE",
                    "tipo_combustible",
                    "Tipo Combustible",
                ),
                self._get_value(
                    item,
                    "PESO_BRUTO",
                    "peso_bruto",
                    "Peso Bruto",
                ),
                self._get_value(
                    item,
                    "APROBACION",
                    "aprobacion",
                    "Aprobación",
                ),
            ]

            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        self._auto_width(ws)

    def _get_value(self, item, *keys):
        if not isinstance(item, dict):
            return None

        for key in keys:
            value = item.get(key)

            if value is not None:
                return value

        return None

    def _style_header(self, cell):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _auto_width(self, ws):
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0

            for cell in column_cells:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))

            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)